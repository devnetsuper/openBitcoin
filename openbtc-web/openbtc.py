from flask import Flask, render_template, request, send_from_directory, redirect, url_for
from openpyxl.styles import colors, PatternFill, Font, Color
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from decimal import Decimal
import requests
import json
import os

TEMP_DIR_VERIFY_TX = 'temp/verify-tx/'
TEMP_DIR_PRICE_HISTORY = 'temp/price-history/'
TEMP_DIR_SENT_RECEIVED = 'temp/sent-received/'

BTC_SYMBOL = 'BTC'
BTC_DECIMAL = 1e8

app = Flask(__name__)

btcusd_history = None

@app.route('/init')
def init():
    exchange_history = requests.get(f'https://api.cryptowat.ch/markets/bitfinex/btcusd/ohlc?&periods=86400')
    content = json.loads(exchange_history.content)
    global btcusd_history
    btcusd_history = content['result']['86400']
    return render_template('index.html')

@app.route('/')
def index():
   return render_template('index.html')
   
def tx_exists(tx_hash):
    is_valid = requests.get(f'https://mempool.space/api/tx/{tx_hash.strip()}')
    
    if is_valid.status_code == 200:
        return True
    
    return False   

@app.route('/verifyaddress', methods=['GET'])
def verify_address():
    address = request.args.get('address')
    is_valid = requests.get(f'https://mempool.space/api/address/{address}')

    if is_valid.status_code == 200:
        return render_template('index.html')

    return "Invalid address", 400
    
@app.route('/verifytx', methods=['GET'])
def verify_tx():
    tx_hash = request.args.get('tx_hash')
    
    if tx_exists(tx_hash):
        return render_template('index.html')
    
    return "Invalid transaction", 400   
    
def is_file_allowed(filename):
    return filename and filename.lower().endswith('.xlsx')
    
def verify_multi_tx(file):
    work_book = load_workbook(file, data_only=True)
    
    for worksheet_idx in range(len(work_book.worksheets)):
        tx_id_idx = None
        bc_url_idx = None
        work_book.active = worksheet_idx
        work_sheet = work_book.active
        #upload_status.set(f'Working with sheet {work_sheet.title}')
       
        for column_idx, column_name in enumerate(work_sheet[1]):
            if column_name.value in ('Transaction Details', 'Transaction ID', 'Transaction Detail'):
                tx_id_idx = column_idx
            elif column_name.value == 'Blockchain URL':
                bc_url_idx = column_idx


        row_count = work_sheet.max_row
       
        for row_idx in range(2, row_count + 1):
            #upload_status.set(f'Working with row #{row_idx}/{row_count}')
            if tx_id_idx is not None and bc_url_idx is not None:
                tx_id = work_sheet[row_idx][tx_id_idx].value
                if tx_exists(tx_id):
                    bc_link = f'https://mempool.space/tx/{tx_id}'
                    work_sheet[row_idx][bc_url_idx].value = '=HYPERLINK("{}", "{}")'.format(bc_link, 'Verified')
                    work_sheet[row_idx][bc_url_idx].fill = PatternFill(start_color='C3ECCB', end_color='C3ECCB', fill_type = 'solid')
                    work_sheet[row_idx][bc_url_idx].font = Font(color='006100')
                else:
                    work_sheet[row_idx][bc_url_idx].value = 'Null'
                    work_sheet[row_idx][bc_url_idx].fill = PatternFill(start_color='F2D3D7', end_color='F2D3D7', fill_type = 'solid')
                    work_sheet[row_idx][bc_url_idx].font = Font(color='9C0039')

    work_book.save(file)                        

def clean_dir(dir_name):
    for _, _, files in os.walk(dir_name):
        for file in files:
            print(f'removing {file}...')
            os.remove(f'{dir_name}{file}')
            
@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    if is_file_allowed(file.filename):
        clean_dir(TEMP_DIR_VERIFY_TX)
        file_path = f'{TEMP_DIR_VERIFY_TX}{file.filename}'
        file.save(file_path)
        verify_multi_tx(file_path)
        return render_template('index.html')
        
    return "Invalid file", 400
    
@app.route('/download-vtx/<filename>', methods=['GET'])
def download_vtx(filename):
    print(f'downloading {filename}...')
    return send_from_directory(directory=os.path.join(app.root_path, TEMP_DIR_VERIFY_TX), path=filename, as_attachment=True)
    

def get_history(market, before, after):
    history = requests.get(f'https://api.cryptowat.ch/markets/{market}/btcusd/ohlc?before={before}&after={after}&periods=86400')
    content = json.loads(history.content)
    return content['result']['86400']


@app.route('/priceHistory', methods=['GET'])
def generate_historical_price_data():
    clean_dir(TEMP_DIR_PRICE_HISTORY)
    
    before = int(request.args.get('before'))
    after = int(request.args.get('after'))
    market = request.args.get('market')

    rows = get_history(market, before + 28800, after)
    
    filename = f'{TEMP_DIR_PRICE_HISTORY}/BTC_price_history_{after}-{before}.xlsx'
    
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.append(['Date', 'Closing Price(USD)'])

    for row in rows:
        date = datetime.fromtimestamp(int(row[0])).strftime('%m/%d/%Y')
        work_sheet.append([date, row[4]])        
    
    work_book.save(filename)
    return render_template('index.html')   

@app.route('/download-hpd/<filename>', methods=['GET'])
def download_hpd(filename):
    print(f'downloading {filename}...')
    return send_from_directory(directory=os.path.join(app.root_path, TEMP_DIR_PRICE_HISTORY), path=filename, as_attachment=True)
    
def get_current_value():
    price = requests.get('https://api.cryptowat.ch/markets/bitfinex/btcusd/price')
    content = json.loads(price.content)
    return Decimal(content['result']['price'])
    
def get_entry_price(tx_time):
    days = tx_time // 86400
    tx_time = days * 86400
    for row in btcusd_history:
        if row[0] == tx_time:
            return Decimal(row[4])
    
    return 1

def get_txs(address, start_date, end_date, include_ugl):
    #startdate = int(datetime.strptime(start_date.get(), "%m/%d/%y").timestamp())
    #enddate = int(datetime.strptime(end_date.get(), "%m/%d/%y").timestamp())
    
    filtered_txs = []

    last_seen_txid = ''
    last_seen_txid = get_txs_chain(address, start_date, end_date, filtered_txs, last_seen_txid, include_ugl)
    while last_seen_txid:
        last_seen_txid = get_txs_chain(address, start_date, end_date, filtered_txs, last_seen_txid, include_ugl)
        
    return filtered_txs

def get_total_value(address, funds):
    total = 0
    for fund in funds:
        if 'prevout' in fund:
            fund = fund['prevout']
        if 'scriptpubkey_address' in fund:
            fund_address = fund['scriptpubkey_address']
            if address == fund_address:
                total += int(fund['value'])
    
    return total
    
def round_2_decimal_places(amount):
    amount *= 100
    amount = round(amount)
    return amount / 100
    
def get_txs_chain(address, start_date, end_date, filtered_txs, last_seen_txid, include_ugl):
    txs = requests.get(f'https://mempool.space/api/address/{address}/txs/chain/{last_seen_txid}')
    
    content = json.loads(txs.content)
    
    if txs.status_code == 400:
        return ''
        
    current_value = get_current_value()
    
    last_tx = ''
    for tx in content:
        tx_time = int(tx['status']['block_time']) - 28800
        
        if tx_time < start_date:
            break

        last_tx = tx['txid']

        if tx_time > end_date:
            continue

        sent = get_total_value(address, tx['vin'])/BTC_DECIMAL
        received = get_total_value(address, tx['vout'])/BTC_DECIMAL
        
        if sent != 0 and received != 0:
            if sent > received:
                sent -= received
                received = 0
            else:
                received -= sent
                sent = 0
            
        tx_data = [datetime.fromtimestamp(tx_time).strftime('%Y-%m-%d %H:%M:%S'),
                   tx['status']['block_height'], tx['txid'], sent, received, BTC_SYMBOL]

        if include_ugl:
            value = sent if sent != 0 else received
            tx_data.append(f'${round_2_decimal_places(current_value * Decimal(value))}')
            entry_price = get_entry_price(tx_time + 57600)
            tx_data.append(f'${round_2_decimal_places(entry_price * Decimal(value))}')
            
        filtered_txs.append(tx_data)

    return last_tx

@app.route('/sentReceived', methods=['GET'])
def generate_sent_received_history():
    #send_received_status.set('Retrieving Transactions...')
    clean_dir(TEMP_DIR_SENT_RECEIVED)
    
    address = request.args.get('address').strip()
    start_date = int(request.args.get('start_date'))
    end_date = int(request.args.get('end_date'))
    include_ugl = request.args.get('include_ugl')
    rows = get_txs(address, start_date, end_date, include_ugl)
    
    filename = f'{TEMP_DIR_SENT_RECEIVED}/BTC_{address}.xlsx'
    
    work_book = Workbook()
    work_sheet = work_book.active
    columns = ['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset']
    
    if include_ugl:
        columns.append('Current USD Value (@currentprice * amount)')
        columns.append('Entry Price Value(USD)')

    work_sheet.append(columns)
    
    for row in rows:
        work_sheet.append(row)
        
    work_book.save(filename)
    
    return render_template('index.html')
    #send_received_status.set(f'Transactions retrieved successfully as of {datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")}')
    
@app.route('/download-srh/<filename>', methods=['GET'])
def download_srh(filename):
    print(f'downloading {filename}...')
    return send_from_directory(directory=os.path.join(app.root_path, TEMP_DIR_SENT_RECEIVED), path=filename, as_attachment=True)
    
if __name__ == '__main__':
   app.run()